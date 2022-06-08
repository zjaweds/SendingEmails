import smtplib
from email.message import EmailMessage
import imghdr
import pyqrcode 
import png
from pyqrcode import QRCode
import openpyxl

EMAIL_ADDRESS = 'emsjamia@gmail.com'
EMAIL_PASSWORD = 'LatEMS@2'
file_name ="DelhiKop Entry Pass"
msg = EmailMessage()
msg['Subject'] = 'Your QR Code for DelhiKop Screening Event'
msg['From'] = EMAIL_ADDRESS

# wb = openpyxl.load_workbook('payment.xlsx')
# ws = wb.active
total_rows = 32

for i in range(2,total_rows):
    wb = openpyxl.load_workbook('payment.xlsx')
    ws = wb.active
    s = ws.cell(row=i, column=2).value 
    name =  ws.cell(row=i, column=12).value
    paymentId = ws.cell(row=i, column=9).value
    ticket_count = ws.cell(row=i, column=5).value
    mobile_numer = ws.cell(row=i, column=11).value
    email = ticket_count = ws.cell(row=i, column=10).value
    url = pyqrcode.create(s)
    qrFile = s
    url.png(s, scale=6)
    msg['To'] = email
    message = f'''
        Name: {name}
    '''
    msg.set_content(message)

    with open(s, 'rb') as f:
        file_data= f.read()
        file_type = imghdr.what(f.name)
        print(file_type)


    msg.add_attachment(file_data, maintype = 'image', subtype=file_type, filename=file_name)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)
    print(str(i-1)+". Email sent to: "+str(email))