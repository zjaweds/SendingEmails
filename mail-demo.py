import smtplib
from email.message import EmailMessage
import imghdr
import pyqrcode 
import png
from pyqrcode import QRCode
import openpyxl

EMAIL_ADDRESS = 'xyz@email.com'
EMAIL_PASSWORD = 'Password'
file_name ="Custom File Name for the Attachment"
# msg = EmailMessage()
# msg['Subject'] = 'Your QR Code for DelhiKop Screening Event'
# msg['From'] = EMAIL_ADDRESS

wb = openpyxl.load_workbook('emailids.xlsx')
ws = wb.active
# total_rows = ws.max_row
# print(total_rows)

for i in range(81,86):
    msg = EmailMessage()
    msg['Subject'] = 'Your QR Code for Planned Event'
    msg['From'] = EMAIL_ADDRESS
    paymentId = s = ws.cell(row=i, column=7).value 
    name =  ws.cell(row=i, column=4).value
    # print(paymentId)
    ticket_count = ws.cell(row=i, column=3).value
    # print(ticket_count)
    mobile_numer = ws.cell(row=i, column=5).value
    # print(mobile_numer)
    # print('_____'+str(i)+'________')
    email  = ws.cell(row=i, column=6).value
    url = pyqrcode.create(s)
    qrFile = s
    url.png(s, scale=6)
    msg['To'] = email
    message = f'''
        Name: {name}

        Payment Id: {paymentId}
        
        Ticket Count: {ticket_count}
        
        Mobile No.: {mobile_numer}
    '''
    msg.set_content(message)

    with open(s, 'rb') as f:
        file_data= f.read()
        file_type = imghdr.what(f.name)
        print(file_type)


    msg.add_attachment(file_data, maintype = 'image', subtype=file_type, filename=file_name)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)
    print(str(i-1)+". Email sent to: "+str(email))
    msg.clear()